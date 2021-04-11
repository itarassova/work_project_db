import json
import requests
from jsonpath_ng import jsonpath
from jsonpath_ng.ext import parser
from openpyxl import Workbook, load_workbook
import time
import logging as log
from compound import Compound
from hazard import Hazard, HazardTypes
from sql import Database
from pubchem import get_hazards

start = time.time()

count_issues_worksheet = 1  

wb = load_workbook(filename='InventoryExportFinal.xlsx')
ws = wb.active

issues_workbook = Workbook()
issues_worksheet = issues_workbook.active

database = Database('Charnwood_inventory_back-up_really_large_number.db')

read_row_index = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    read_row_index += 1
    try:
        name = row[0]
        cas = row[1]
        compound = Compound(cas, name)
        if read_row_index % 100 == 0:
            log.info("Reading record {}", str(read_row_index))
    except Exception as e:
        log.error(e, exc_info=True)
    compound_in_db = database.get_reagent_id_from_db(compound)
    if not compound_in_db:
        try:
            hazards, _ = get_hazards(compound)
            database.insert_compound_hazard(compound, hazards)
        except Exception as e:
            log.error(e, exc_info=True)
            issues_worksheet.append([str(compound.cas), str(e)])
            count_issues_worksheet += 1 

issues_workbook.save(filename="db_issues_output.xlsx")

end = time.time()
log.info("Execution took: %s", str(end-start))
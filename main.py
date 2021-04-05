#!/bin/python3

import json
import requests
from jsonpath_ng import jsonpath
from jsonpath_ng.ext import parser
from openpyxl import Workbook, load_workbook
import time
import logging as log
from quantity import Quantity
from compound import Compound
from hazard import Hazard, HazardTypes
from sql import Cache

def get_room_from_location(location):
    location_split = location.split('>')[1]
    substring = ' - '
    if substring in location_split:
        room = location_split.split('-')[1]
    else:
        room = location_split.split()[1]
    return room



def is_explosive(list : list[Hazard]) -> bool:
    not_explosive_290 = 'H290'
    not_explosive_281 = 'H281'

    if list:
        hazard_codes_list = [(warning_line[:4]) for warning_line in list]
        for hazard_code in hazard_codes_list:
            if not_explosive_290.casefold() != hazard_code.casefold() and not_explosive_281.casefold() != hazard_code.casefold():
                return True
    return False

start = time.time()

wb = load_workbook(filename='InventoryExportFinal.xlsx')
ws = wb.active

export_workbook = Workbook()
export_worksheet = export_workbook.active

issues_workbook = Workbook()
issues_worksheet = issues_workbook.active

count_issues_worksheet = 1    

# substances = {"1234-56-78": {"Room 1": "15 ml", "Room 2": "30 ml"}, "9876-5-32" {"Room 1": "15 g", "Room 2": "30 g"}}
substances = {}

cache = Cache('Charnwood_inventory_back-up.db')

read_row_index = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    read_row_index += 1
    try:
        name = row[0]
        cas = row[1]
        location = row[7]
        amount = row[2]
        unit = row[3]
        compound = Compound(cas, name)
        if read_row_index % 100 == 0:
            log.info("Reading record {}", str(read_row_index))
        room_location = get_room_from_location(location)
        amount = Quantity(amount, unit)

        locations = substances.get(compound, {})
        amount_in_location = locations.get(room_location)
        if not amount_in_location:
            locations[room_location] = amount
        else:
            amount_in_location += amount
        substances[compound] = locations
    except Exception as e:
        log.error(e, exc_info=True)
        issues_worksheet.append(row)
        count_issues_worksheet += 1


row_number = 2
for compound in substances:
    try:
        compound_in_db = cache.get_compound_from_db(compound)
        if not compound_in_db:
            locations = substances[compound]
            cache.insert_compound_hazard(compound) 
        try:
            physical_warning_lines, health_warning_lines, environmental_warning_lines, other_hazards, msds_url, further_information = get_hazards(
                compound)
            explosive = "Yes" if is_explosive(physical_warning_lines) else "No"
            physical_hazards = '\n'.join(physical_warning_lines)
            health_hazards = '\n'.join(health_warning_lines)
            environmental_hazards = '\n'.join(environmental_warning_lines)
            other_hazards = '\n'.join(other_hazards)
            further_information = '\n'.join(further_information)
        except:
            explosive = ""
            physical_hazards = ""
            health_hazards = ""
            environmental_hazards = ""
            other_hazards = ""
            further_information = ""
        for location in locations:
            row_number += 1
            if row_number % 50 == 0:
                log.info("Writing record {}", str(row_number))
            amount = locations[location]
            export_worksheet['C' + str(row_number)] = str(compound.name)
            #export_worksheet['B'+ str(row_number)] = str(compound.cas)
            export_worksheet['H' + str(row_number)] = str(physical_hazards)
            export_worksheet['I' + str(row_number)] = str(health_hazards)
            export_worksheet['J' + str(row_number)
                             ] = str(environmental_hazards)
            #export_worksheet['F'+ str(row_number)] = str(other_hazards)
            export_worksheet['G' + str(row_number)] = str(msds_url)
            export_worksheet['B' + str(row_number)] = str(location)
            export_worksheet['K' + str(row_number)] = str(explosive)
            export_worksheet['D' + str(row_number)] = str(amount)
            export_worksheet['E' + str(row_number)] = str('Wet chemistry')
            export_worksheet['A' + str(row_number)] = str('BCN - Stuart Adams')
            export_worksheet['M' + str(row_number)] = str(further_information)
            export_worksheet['F' + str(row_number)] = str('Yes')

    except Exception as e:
        log.error(e, exc_info=True)
        issues_worksheet.append(
            [str(compound.cas), str(e)])
        count_issues_worksheet += 1


export_workbook.save(filename="Biocity_output.xlsx")
issues_workbook.save(filename="issues_output.xlsx")
end = time.time()
log.info("Execution took: {}", str(end-start))

#!/bin/python3

import json
import requests
from jsonpath_ng import jsonpath
from jsonpath_ng.ext import parser
from openpyxl import Workbook, load_workbook
import time
import logging as log


start = time.time()

wb = load_workbook(filename='inventoryexport_trial.xlsx')
ws = wb.active

export_workbook = Workbook()
export_worksheet = export_workbook.active

row = 2

def get_room_from_location(location):
    location_split = location.split('>')[1]
    substring = ' - '
    if substring in location_split:
        room = location_split.split('-')[1]
    else:
        room = location_split.split()[1]
    return room

def convert_quantity_units(amount, unit):
    mass_unit_mg = 'mg'
    mass_unit_kg = 'Kg'
    if unit == mass_unit_mg:
        converted_amount = int(amount) * 0.001
    if unit == mass_unit_kg:
        converted_amount = int(amount) * 1000
    volume_unit_L = 'L'
    if unit == volume_unit_L:
        converted_amount = int(amount) * 1000
    return converted_amount


for cell in ws['B']:
    try:
        cas = cell.value
        cid_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{}/cids/JSON".format(cas)
        cid_resp = requests.get(url=cid_url)
        cid = cid_resp.json()["IdentifierList"]["CID"][0]
        log.info(cid)
        msds_url = "https://pubchem.ncbi.nlm.nih.gov/compound/{}#datasheet=LCSS".format(cas)
        compound_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{}/JSON".format(cid)
        resp = requests.get(url=compound_url)
        data = resp.json()
        filter_expression = parser.parse('$.Record.Section[?(@.TOCHeading=="Safety and Hazards")].Section[0].Section[0].Information[2].Value.StringWithMarkup[*].String')
        health_warning_lines = []
        physical_warning_lines = []
        environmental_warning_lines = []
        other_hazards = []
        for warning_line in filter_expression.find(data):
            log.info(warning_line.value)        
            if warning_line.value[1] == '2':
                physical_warning_lines.append(warning_line.value)
            elif warning_line.value[1] == '3':
                health_warning_lines.append(warning_line.value)
            elif warning_line.value[1] == '4':
                environmental_warning_lines.append(warning_line.value)
            else:
                other_hazards.append(warning_line.value)

        physical_hazards = '\n'.join(physical_warning_lines)
        health_hazards = '\n'.join(health_warning_lines)
        environmental_hazards = '\n'.join(environmental_warning_lines)
        other_hazards = '\n'.join(other_hazards)

        export_worksheet['A'+ str(row)] = str(cas)
        export_worksheet['B'+ str(row)] = str(physical_hazards)
        export_worksheet['C'+ str(row)] = str(health_hazards)
        export_worksheet['D'+ str(row)] = str(environmental_hazards)
        export_worksheet['E'+ str(row)] = str(other_hazards)
        export_worksheet['F'+ str(row)] = str(msds_url)
        export_worksheet['G'+ str(row)] = str(ws['A'+str(row)].value)
        room_location = get_room_from_location(ws['H'+str(row)].value)
        export_worksheet['H'+ str(row)] = str(room_location)
        compound_unit = ['B'+ str(row)]
        row+=1

    except Exception as e:
        print(str(cas) + ' didn\'t work')
        log.error(e, exc_info=True)

export_workbook.save(filename="test_output.xlsx")
end = time.time()
print(end-start)

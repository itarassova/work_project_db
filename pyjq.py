#!/bin/python3

import json
import requests
from jsonpath_ng import jsonpath, parse
from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='inventoryexport_trial.xlsx')
ws = wb.active

export_workbook = Workbook()
export_worksheet = export_workbook.active

row = 2

list_cas = []

for col in ws['B']:
    list_cas.append(col.value)

for cas in list_cas:
    try:
        cid_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{}/cids/JSON".format(cas)
        cid_resp = requests.get(url=cid_url)
        cid = cid_resp.json()["IdentifierList"]["CID"][0]
        print(cid)
        
        compound_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{}/JSON".format(cid)
        resp = requests.get(url=compound_url)
        data = resp.json()
        for i in range (7,12):
            filter_expression = parse('$.Record.Section[{}].Section[0].Section[0].Information[2].Value.StringWithMarkup[*].String'.format(i))
            warning_lines = []
            for warning_line in filter_expression.find(data):
                print(warning_line.value)
                warning_lines.append(warning_line.value)
            join_lines = ' '.join(warning_lines)
            print(join_lines)

        export_worksheet['A'+ str(row)] = str(cas)
        export_worksheet['B'+ str(row)] = join_lines
        export_workbook.save(filename="test_output.xlsx")
        row+=1
        
		
    except:
        print(str(cas) + ' didn\'t work')
        
